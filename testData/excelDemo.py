import openpyxl
Dict = {}
# reading from sheet
book = openpyxl.load_workbook("../PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2).value = "Kranthi"
print(sheet.cell(row=2, column=2).value)

print(sheet.max_column)
print(sheet.max_row)

print(sheet['A5'].value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i,column=1).value == "Testcase2":
        for j in range(1, sheet.max_column+1):
            # print(sheet.cell(row=i, column=j).value)
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)