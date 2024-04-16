import openpyxl

class HomePageData:

    test_HomePage_data = [{"firstname": "Kranthi", "lastname": "Panda", "gender": "Male"},
                            {"firstname": "Sagar", "lastname": "Mittapelly", "gender": "Male"},
                          {"firstname": "Ravi Teja", "lastname": "Korati", "gender": "Male"},
                          {"firstname": "Siva", "lastname": "Lakshmi", "gender": "Female"},
                          {"firstname": "Venkatesh", "lastname": "Godishala", "gender": "Male"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("../PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == "Testcase2":
                for j in range(1, sheet.max_column + 1):  # to get columns
                    # print(sheet.cell(row=i, column=j).value)
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]